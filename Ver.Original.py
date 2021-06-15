import discord
from discord.ext import commands
from discord.utils import get
import urllib.request
import requests
from bs4 import BeautifulSoup
import ctx
import asyncio
import pandas as pd
import openpyxl
from datetime import timedelta
from datetime import date
import datetime
from openpyxl import Workbook
import permission
import schedule
import time

     
app = discord.Client()

wb = Workbook()
#
# reactions = ["✅"]
#
# @app.event
# async def on_reaction_add(reaction, user):
#     if str(reaction.emoji) == "✅":
#         print("1")
#         Role = discord.utils.get(reaction.server.roles, name="회원")
#         await user.add_role(reaction.user, Role)
#         role2 = discord.utils.get(reaction.server.roles, name="인증필요계정")
#         await user.delete_role(reaction.user, role2)

@app.event
async def on_voice_state_update(member, before, after):
    channel1 = app.get_channel(737942276684120085)
    embed99=discord.Embed(title="", description=(str(after.channel))+"에 접속하였습니다", color=0x00ff00)
    embed99.set_author(name=member,icon_url=member.avatar_url)
    await channel1.send(embed = embed99)
    username = str(member)
    guild = app.get_guild(589828103635730452)              #채널아이디
    ch = guild.get_channel(774353602428272640)             #입력 채널 아이디
    category = guild.get_channel(692942116833460266)       #카테고리 아이디
    if after.channel == ch:
        channel = await guild.create_voice_channel(
            name=member.display_name+"`s Room",
            category=category,
            user_limit=5,
            bitrate=256000
        )
        await member.move_to(channel)
        await channel.set_permissions(member, manage_channels=True)
    if not before.channel.voice_states and before.channel != ch:
        await before.channel.delete()


@app.event
async def on_ready():
    print(app.user.name)
    print(app.user.id)


@app.event
async def on_message(message):
    if message.author.bot:
        return


    if message.content.startswith('!추천인'):
            file = openpyxl.load_workbook("추천인.xlsx")
            sheet = file.active
            learn = message.content.split(" ")
            for i in range(1, 1000):           #1~1000 까지중 C에 같은내용이 있다면 D에 숫자 1을 추가로 부여한다.
                if str(sheet["C" + str(i)].value) == str(message.author.id):
                     sheet["D" + str(i)].value = int(sheet["D" + str(i)].value) + 1
                     if int(sheet["D" + str(i)].value) == 2:        #D에 숫자가 2 이상일시 메세지를 전송한다.
                         await message.channel.send("추천은 한번만 할 수 있습니다.")
                     break
                if sheet["C"+str(i)].value == "_":             #C의 시트가 _ 일시 빈 시트로 판단한다.
                    sheet["A" + str(i)].value = learn[1]
                    sheet["B" + str(i)].value = learn[2]        #C의 시트에 메세지의 주인의 ID를 적고 D에 1을 부여한다. D는 추천을 한 횟수이다.
                    sheet["C" + str(i)].value = str(message.author.id)
                    sheet["D" + str(i)].value = 1
                    await message.channel.send("추천이 완료되었습니다.")
                    break
                
                if sheet["F"+str(i)].value == learn[1]:          #F시트에 스플릿한 내용의 1번과 같은 내용이 있다면
                    sheet["G" + str(i)].value = int(sheet["G" + str(i)].value) + 1      #G시트에 1을 추가한다.      G시트는 추천을 받은 횟수이다.
                    if sheet["A"+str(i)].value == "_":
                         sheet["A" + str(i)].value = learn[1]
                         sheet["B" + str(i)].value = learn[2]
                         sheet["C" + str(i)].value = str(message.author.id)            #C시트에 메세지 주인의 ID를 입력한다.
                         await message.channel.send("추천이 완료되었습니다.")
                         break


                for i in range(1, 1000):
                    if sheet["A"+str(i)].value == learn[1]:
                        sheet["F" + str(i)].value = learn[1]
                        break
                        await message.channel.send("추천이 완료되었습니다.")
                        

               
            file.save("추천인.xlsx")
            await app.get_channel(int(701381051137654845)).send(message.content)
            await app.get_channel(int(701381051137654845)).send(message.author.id)
            await app.get_channel(int(701381051137654845)).send("-----------next user-------------")

            

    if message.content.startswith("!DM"):
        ids = [member.id for member in message.guild.members]
        for ary in ids:
            try:
                _var = message.guild.get_member(ary)
                channel = await _var.create_dm()
                await channel.send(message.content[4:])
            except discord.errors.Forbidden:
                    pass
    

    if message.content.startswith('!롤'):
        #Name = message.content.split(" ")
        #Name1 = Name[1]
        #location = message.content[3:len(message.content)]

        Name = message.content[3:len(message.content)]

        req = requests.get('http://op.gg/summoner/userName='+Name)
        html = req.text
        soup = BeautifulSoup (html, 'html.parser')
        

        ####################### 닉네임 부분 ###################################

        Name1 = soup.find_all('span', {'class': 'Name'})
        Name2 = str(Name1[0])[str(Name1[0]).find('me">') +4:str(Name1[0]).find('</span>')]
        print(Name2)

        ####################### 프로필 인장 부분 ##############################

        Image1 = soup.find_all('div', {'class': 'borderImage'})
        Image2 = str(Image1[0])[str(Image1[0]).find('img">') +4:str(Image1[0]).find('</div>')]
        print(Image2)

        ####################### 모스트 부분 ###################################

        Most1 = soup.find_all('div', {'class': "ChampionInfo"})
        Most2 = str(Most1[0])[str(Most1[0]).find('nt">') +4:str(Most1[0]).find('</div>')] ##손질중##
        print(Most2)


        
        ######################### 랭크 부분 #########################
        
        Rank1 = soup.find_all('div', {'class': 'TierRank'})
        Rank2 = str(Rank1[0])[str(Rank1[0]).find('nk">') +4:str(Rank1[0]).find('</div>')]
        print(Rank2)

        ######################### 점수 부분 #########################
        
        LP1 = soup.find_all('span', {'class': 'LeaguePoints'})
        LP2 = str(LP1[0])[str(LP1[0]).find('">')+2 :str(LP1[0]).find('</sp')]
        LP3 = LP2.strip()
        print(LP3)
        
        ######################### 승리,패배,승률 부분#########################
        
        win1 = soup.find_all('span', {'class': 'WinLose'})
        win2 = str(win1[0])[str(win1[0]).find('ns">') + 4:str(win1[0]).find('</sp')]
        print(win2)
                              
        lose1 = soup.find_all('span', {'class': 'losses'})
        lose2 = str(lose1[0])[str(lose1[0]).find('es">') + 4:str(lose1[0]).find('</sp')]
        print(lose2)

        ratio1 = soup.find_all('span', {'class': 'winratio'})
        ratio2 = str(ratio1[0])[str(ratio1[0]).find('io">') + 4:str(ratio1[0]).find('</sp')]
        print(ratio2)

        ###########################가장 최근 전적##################################

        Game1 = soup.find_all('div', {'class': 'GameType'})
        Game2 = str(Game1[0])[str(Game1[0]).find('ga">') + 4:str(Game1[0]).find('</div')]
        print(Game2)

        Result1 = soup.find_all('div', {'class': 'GameResult'})
        Result2 = str(Result1[0])[str(Result1[0]).find('lt">') + 4:str(Result1[0]).find('</div')]
        print(Result2)

        ########################## 총 정리 ######################################

        
        Image3 = Image2.replace('  ','  ')
        Result3 = Result2.replace('GameResult',' ')
        Game3 = Game2.replace('GameType','게임종류')
        win3 = win2.replace('W','승')
        lose3 = lose2.replace('L','패')
        ratio3 = ratio2. replace('Win Ratio', '승률')
        Rank3 = Rank2. replace('TierRank', '티어')
        LP4 = LP3. replace('LeaguePoints', 'LP')
        Most3 = Most2. replace('ChampionInfo', 'Most')
        Name3 = Name2. replace('Name', '닉네임')
        print(win3+'/'+lose3+'/'+ratio3)

       ############################출력부분######################################
        
        embed=discord.Embed(title="LOL 전적검색", description="", color=0x00ff00)
        embed.set_author(name=message.author.name,icon_url=message.author.avatar_url)
        embed.add_field(name="닉네임", value= (Name3), inline=True)
        embed.add_field(name="솔랭티어", value= (Rank3+' / '+LP4) ,inline=True)
        embed.add_field(name="승,패", value= (win3+' / '+lose3+' / '+ratio3), inline=True)
        embed.set_footer(text="롤토피아 롤 전적검색")
        embed.set_image(url="https://search.pstatic.net/common/?src=http%3A%2F%2Fblogfiles.naver.net%2F20160412_229%2Fhkkimhan_14604701270821RHjm_JPEG%2Fmaxresdefault.jpg&type=sc960_832")
        await message.channel.send(embed=embed)
        role = ""
        rolename = Rank3.split(' ')
        rank = rolename[0]
        role = discord.utils.get(message.guild.roles, name=rank)
        await message.author.add_roles(role)
        print(role)
        req.close()
        soup.clear()
        
        #############################디스코드 부분######################

    if message.content.startswith ("!스크림추가"):
        embed4 = discord.Embed(title="---스크림 추가양식---", description="", color=0x00ff00)
        embed4.set_author(name=message.author.name,icon_url=message.author.avatar_url)
        embed4.add_field(name="!팀명:", value= "팀명을 적어주세요.", inline=False)
        embed4.add_field(name="날짜:", value= "날짜를 적어주세요.", inline=False)
        embed4.add_field(name="시간:", value= "시간을 적어주세요.", inline=False)
        embed4.add_field(name="평균티어:", value= "팀 내 평균티어를 적어주세요.", inline=False)
        embed4.add_field(name="팀장 롤닉네임:", value= "롤 닉네임을 적어주세요.", inline=False)
        embed4.set_footer(text="꼭 양식을 포함하여 적어주세요.")
        await message.channel.send(embed=embed4)
                               
    if message.content.startswith ("!팀명:"):
       embed2 = discord.Embed(title="일정추가가 완료되었습니다.", description="", color=0xFF0000)
       embed2.set_author(name=message.author.name,icon_url=message.author.avatar_url)
       embed2.add_field(name="일청추가 완료", value= "본인의 팀과 경기를 희망하는 팀 혹은 관리자의 연락이 갈 수 있습니다.", inline=True)
       await message.channel.send(embed=embed2)
       await app.get_channel(int(689935648098091174)).send(message.content)


    if message.content.startswith ("!신고"):
       embed5 = discord.Embed(title="--------양식--------", description="", color=0xFF0000)
       embed5.add_field(name="!가해자 닉네임", value= "가해자 디스코드 닉네임을 기입해주세요.", inline=False)
       embed5.add_field(name="신고사유:", value= "신고사유를 기입해주세요. 증거도 있을시 증거는 일분관리자 에게 제출해주세요.", inline=False)
       embed5.add_field(name="신고자 닉네임", value= "본인의 디스코드 닉네임을 기입해주세요.", inline=False)
       embed5.set_footer(text="꼭 양식을 포함하여 적어주세요.")
       await message.channel.send(embed=embed5)
       await message.delete()
                               
    if message.content.startswith ("!가해자"):
       embed13 = discord.Embed(title="신고가 완료되었습니다.", description="관리자에게서 연락이 가실 수 있습니다.", color=0xFF0000)
       await message.channel.send(embed=embed13)
       await app.get_channel(int(689936339696877568)).send(message.content)
       await message.delete()
                
    if message.author.id == (651495978075029572) and (364390750206558214):
        if message.content.startswith("!재제 "):
            await app.get_channel(826742895528640542).send("경고 처리까지 약 1분의 시간이 소요됩니다.")
            message1 = message.content[4:len(message.content)]
            message2 = str(message1)
            message3 = message2.split(" ")
            messagenick = message3[0]
            messageReason = message3[1]
            messageContents = message3[2]
            messageManager = message.author.id
            print(messagenick)
            print(message3[1])
            print(message3[2])
            print(messageManager)
            embed4 = discord.Embed(title="재제 알림", description="", color=0x00ff00)
            embed4.add_field(name="위 반 인:", value="@"+str(messagenick), inline=False)
            embed4.add_field(name="제재사유:", value=str(messageReason), inline=False)
            embed4.add_field(name="제재내용:", value=str(messageContents), inline=False)
            embed4.add_field(name="처 리 자:", value="@"+str(messageManager), inline=False)
            await app.get_channel(723152658667667496).send("**재제 알림**\n"
                                                           "%s"%messagenick+"\n"
                                                           "제재사유: "+messageReason+"\n"
                                                           "제재내용: "+messageContents+"\n"
                                                           "처 리 자: <@%s>"%message.author.id)
            wb = Workbook()
            file = openpyxl.load_workbook("제재목록.xlsx")
            sheet = file.active
            for i in range(1, 5001):  # 1~1000 까지중 C에 같은내용이 있다면 D에 숫자 1을 추가로 부여한다.
                if str(sheet["A" + str(i)].value) == str(messagenick):
                    sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                    file.save("제재목록.xlsx")
                    file.close()
                    if int(sheet["B" + str(i)].value) == 3:
                        await app.get_channel(826742895528640542).send("경고가 3회 이상 누적된 유저입니다.(현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                        file.save("제재목록.xlsx")
                        file.close()
                        break
                    else:
                        if int(sheet["B" + str(i)].value) > 3:
                            await app.get_channel(826742895528640542).send("경고가 3회 이상 누적된 유저입니다.(현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                            file.save("제재목록.xlsx")
                            file.close()
                        else:
                            if int(sheet["B" + str(i)].value) < 3:
                                await app.get_channel(826742895528640542).send("경고가 3회 이상 누적된 유저입니다. (현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                            file.save("제재목록.xlsx")
                            file.close()
                            break
                    break
                if sheet["A" + str(i)].value == "_":  # C의 시트가 _ 일시 빈 시트로 판단한다.
                    sheet["A" + str(i)].value = str(messagenick)
                    sheet["B" + str(i)].value = "1"  # C의 시트에 메세지의 주인의 ID를 적고 D에 1을 부여한다. D는 추천을 한 횟수이다.
                    sheet["C" + str(i)].value = str(message.author.id)
                    await app.get_channel(826742895528640542).send("기록이 완료되었습니다. (현재 누적수: 경고"+str(sheet["B" + str(i)].value)+"회)")
                    file.save("제재목록.xlsx")
                    file.close()
                    break
        if message.content.startswith("!가재제"):
            await app.get_channel(826742895528640542).send("가경고 처리까지 약 1분의 시간이 소요됩니다.")
            message1 = message.content[5:len(message.content)]
            message2 = str(message1)
            message3 = message2.split(" ")
            messagenick = message3[0]
            messageReason = message3[1]
            messageContents = message3[2]
            messageDay = date.today()
            week = timedelta(weeks=2)
            print(messageDay)
            messageManager = message.author.id
            print(messagenick)
            print(message3[1])
            print(message3[2])
            print(messageManager)
            await app.get_channel(723152658667667496).send("**재제 알림**\n"
                                                            "%s" % messagenick + "\n"
                                                            "제재사유: " + messageReason + "\n"
                                                            "제재내용: " + messageContents + "\n"
                                                            "제재기간: " +str(messageDay)+ " ~ " +str(date.today()+week)+"\n"
                                                            "처 리 자: <@%s>" % message.author.id+"\n"
                                                            "`(가경고는 2주간 같은사유로 재적발돼지 않을 경우 사라지며, 3회 누적시 영구밴 조치됩니다.)`")
            wb = Workbook()
            file = openpyxl.load_workbook("가경고목록.xlsx")
            sheet = file.active
            for i in range(1, 5001):  # 1~1000 까지중 C에 같은내용이 있다면 D에 숫자 1을 추가로 부여한다.
                if str(sheet["A" + str(i)].value) == str(messagenick):
                    sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 2
                    file.save("가경고목록.xlsx")
                    file.close()
                    if int(sheet["B" + str(i)].value) == 4:
                        await app.get_channel(826742895528640542).send("경고가 3회 이상 누적된 유저입니다.(현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                        file.save("가경고목록.xlsx")
                        file.close()
                        break
                    else:
                        if int(sheet["B" + str(i)].value) > 3:
                            await app.get_channel(826742895528640542).send("경고가 3회 이상 누적된 유저입니다.(현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                            file.save("가경고목록.xlsx")
                            file.close()
                        else:
                            if int(sheet["B" + str(i)].value) < 3:
                                await app.get_channel(826742895528640542).send("가경고가 정상 처리되었습니다.(현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                            file.save("가경고목록.xlsx")
                            file.close()
                            break
                    break
                if sheet["A" + str(i)].value == "_":  # C의 시트가 _ 일시 빈 시트로 판단한다.
                    sheet["A" + str(i)].value = str(messagenick)
                    sheet["B" + str(i)].value = "2"  # C의 시트에 메세지의 주인의 ID를 적고 D에 1을 부여한다. D는 추천을 한 횟수이다.
                    sheet["C" + str(i)].value = str(date.today()+week)
                    sheet["D" + str(i)].value = str(message.author.id)
                    await app.get_channel(826742895528640542).send("기록이 완료되었습니다. (현재 누적수: 경고"+str(sheet["B" + str(i)].value)+"회)")
                    file.save("가경고목록.xlsx")
                    file.close()
                    break

        if message.content.startswith("!가경고해제"):
            await app.get_channel(826742895528640542).send("경고 해제까지 약 30초가 소요됩니다.")
            wb = Workbook()
            file = openpyxl.load_workbook("가경고목록.xlsx")
            sheet = file.active
            id1 = message.content[7:len(message.content)]
            print(id1)
            id2 = str(id1)
            for i in range(1, 5001):
                print(i)
                if str(sheet["A" + str(i)].value) == id2:
                    if int(sheet["B" + str(i)].value) == 2:
                        sheet["A" + str(i)].value = "_"
                        sheet["B" + str(i)].value = " "
                        await app.get_channel(826742895528640542).send(
                            "가경고해제 완료 (현재 누적수: 경고0회)")
                        file.save("가경고목록.xlsx")
                        file.close()
                        break
                    else:
                        if int(sheet["B" + str(i)].value) != 2:
                            sheet["B" + str(i)].value = sheet["B" + str(i)].value - 2
                            await app.get_channel(826742895528640542).send(
                                "가경고해제 완료 (현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                            file.save("가경고목록.xlsx")
                            file.close()
                            break
                else:
                    if i == 5000:
                        await app.get_channel(826742895528640542).send("가경고를 부여받지 않은 유저입니다.")

        if message.content.startswith("!경고해제"):
            await app.get_channel(826742895528640542).send("경고 해제까지 약 30초가 소요됩니다.")
            wb = Workbook()
            file = openpyxl.load_workbook("제재목록.xlsx")
            sheet = file.active
            id1 = message.content[6:len(message.content)]
            print(id1)
            id2 = str(id1)
            for i in range(1, 5001):
                print(i)
                if str(sheet["A" + str(i)].value) == id2:
                    if int(sheet["B" + str(i)].value) == 1:
                        sheet["A" + str(i)].value = "_"
                        sheet["B" + str(i)].value = " "
                        await app.get_channel(826742895528640542).send(
                            "경고해제 완료 (현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                        file.save("제재목록.xlsx")
                        file.close()
                        break
                    else:
                        if int(sheet["B" + str(i)].value) != 1:
                            sheet["B" + str(i)].value = sheet["B" + str(i)].value - 1
                            await app.get_channel(826742895528640542).send(
                                "경고해제 완료 (현재 누적수: 경고" + str(sheet["B" + str(i)].value) + "회)")
                            file.save("제재목록.xlsx")
                            file.close()
                            break
                else:
                    if i == 5000:
                        await app.get_channel(826742895528640542).send("경고를 부여받지 않은 유저입니다.")
        
        
    if message.content.startswith ("!명령어"):
        embed7 = discord.Embed(title="명령어", description="롤토피아 삐약이봇의 명령어입니다.", color=0xFF0000)
        embed7.set_author(name=message.author.name,icon_url=message.author.avatar_url)
        embed7.add_field(name="!스크림추가:", value= "스크림 일정을 추가하실 수 있습니다.", inline=True)
        embed7.add_field(name="!신고:", value= "신고 및 문의방에서만 사용 가능합니다. 신고 혹은 문의사항이 있을때만 사용해주세요..", inline=True)
        embed7.add_field(name="!롤 롤닉네임:", value= "티어인증 방에서 명령어를 사용해 본인의 티어에 맞는 권한을 받으실 수 있습니다.", inline=True)
        await message.channel.send(embed=embed7)



app.run('NjkwNTQyMzgzMTgzNjkxODE3.XnS7tQ.Zw-0AgXIy4a3zEM-bY-FqCi9aRc')


def warning():
    wb = Workbook()
    file = openpyxl.load_workbook("가경고목록.xlsx")
    sheet = file.active
    day = timedelta(days=1)
    for i in range(1, 5001):
        if str(sheet["C" + str(i)].value) == str(date.today() - day):
            sheet["A" + str(i)].value = "_"
            sheet["B" + str(i)].value = "\0"
            sheet["C" + str(i)].value = "\0"
            sheet["D" + str(i)].value = "\0"

schedule.every().day.at("00:10").do(warning)

while True:
    schedule.run_pending()
    time.sleep(0.1)